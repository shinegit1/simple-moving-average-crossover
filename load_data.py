import openpyxl
import mysql.connector

# ------- Establish a MySQL connection -------
my_database =mysql.connector.connect(host="localhost", user="root", password="bright", db="assignment")

# Get the cursor, which is used to traverse the database, line by line
mycursor =my_database.cursor()

# ------- Open the workbook and define the worksheet --------
work_book =openpyxl.load_workbook("HINDALCO_1D.xlsx")
work_sheet =work_book["HINDALCO"]

all_rows =work_sheet.iter_rows(min_row=2,max_row=work_sheet.max_row, max_col=work_sheet.max_column,values_only=True)

# Create the INSERT INTO sql query
query ="insert into hindalco_1d(datetime,close,high,low,open,volume,instrument) values(%s,%s,%s,%s,%s,%s,%s)"

# Assign values from each row
values =[]
for rows in all_rows:
    values.append(rows)

# Execute sql Query
mycursor.executemany(query,values)

# Close the cursor
mycursor.close()

# Commit the transaction
my_database.commit()

# Close the database connection
my_database.close()

print("All done.")